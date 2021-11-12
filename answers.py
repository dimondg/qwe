#Укорачиваем старый фильм
def get_result(name):
    import sqlite3
    con = sqlite3.connect(name)
    cur = con.cursor()
    result = cur.execute('''UPDATE films 
    SET duration = int(duration) / 3
    WHERE year = 1973''')
    con.commit()
    con.close()
#Длинная старая фантастика
def get_result(name):
    import sqlite3
    con = sqlite3.connect(name)
    cur = con.cursor()
    result = cur.execute('''DELETE from films where year < 2000 AND duration > 90 
    AND genre=(SELECT id FROM genres WHERE title = 'фантастика')''').fetchall()
    con.commit()
    con.close()
#Длинные комедии
import sqlite3
f1 = input()
con = sqlite3.connect(f1)
cur = con.cursor()
result = cur.execute("""SELECT title FROM Films WHERE (genre=(SELECT id FROM genres
            WHERE title = 'комедия')) AND duration >= 60""").fetchall()
for elem in result:
    print(elem[0])
con.close()
#Список песен
import sqlite3
nnn = input()
con = sqlite3.connect("music_db.sqlite")
cur = con.cursor()
result = cur.execute("""SELECT DISTINCT Track.Name 
FROM Track
LEFT JOIN album ON track.albumId = album.albumId 
LEFT JOIN artist ON artist.ArtistId = album.ArtistId 
WHERE artist.Name = ? 
ORDER BY track.Name""", (nnn,)).fetchall()
for elem in result:
    print(elem[0])
con.close()
#короткие фильмы
import sqlite3
f1 = input()
con = sqlite3.connect(f1)
cur = con.cursor()
result = cur.execute('''SELECT title FROM films WHERE (duration <= 85)''').fetchall()
for elem in result:
    print(elem[0])

con.close()
#список исполнителей
import sqlite3

a = []
genre = input()
db = sqlite3.connect("music_db.sqlite")
cur = db.cursor()
result = cur.execute(f"""SELECT DISTINCT
  a.name
FROM 
  genre g,
  track t,
  album al,
  artist a
WHERE
  t.genreid = g.genreid 
AND
  t.albumid = al.albumid
AND
  al.artistid = a.artistid
AND
  g.name = '{genre}'
ORDER BY a.name;""")
for i in result:
    if i[0] not in a:
        a.append(i[0])
for i in a:
    print(i)
db.close()

#  Выгодная покупка
import csv
with open('wares.csv', encoding="utf8") as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    expensive = sorted([i for i in reader if int(i[1]) <= 1000], key=lambda x: int(x[1]))
result, k, max_number = [], 1000, 10
for pair in expensive:
    name, price = pair[0], int(pair[1])
    if k >= int(price):
        number_to_buy = k // int(price)
        if number_to_buy > max_number:
            number_to_buy = max_number
        result += [name for _ in range(number_to_buy)]
        k -= number_to_buy * int(price)
if not expensive:
    print('error')
else:
    print(', '.join(result))

# Выпускники
import csv

ch = int(input())
with open('vps.csv', encoding="utf8") as csvfile:
    reader = csv.reader(csvfile, delimiter=';', quotechar='"')
    for i in reader:
        if i[-2] != 'соответствует в %':
            if int(i[-2]) >= ch:
                print(i[0])

#  primo de plantis
import sys
import csv

data = []
s = ['nomen', 'definitio', 'pluma', 'Russian nomen', 'familia', 'Russian nomen familia']
for line in sys.stdin:
    k = {}
    spis = line.rstrip('\n').split('\t')
    for i in spis:
        k[s[spis.index(i)]] = i
    data.append(k)
with open('plantis.csv', encoding="utf8", mode='w') as f:
    writer = csv.DictWriter(
        f, fieldnames=list(data[0].keys()),
        delimiter=";", lineterminator='\r\n')
    writer.writeheader()
    for d in data:
        writer.writerow(d)

#  Скидки
import csv

with open('wares.csv', encoding="utf8") as csvfile:
    reader = csv.DictReader(csvfile, delimiter=';')
    for a in reader:
        if int(a['Old price']) > int(a['New price']):
            print(a['Name'])

#  Результат олипиады
import csv
d = {}
s, cl = input().split()
with open('rez.csv', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for i in list(reader)[1:]:
        info = i[2].split('-')
        name = i[1].split()[3]
        score = int(i[-1])
        if int(s) == int(info[2]) and int(cl) == int(info[3]):
            if str(score) not in d:
                d[str(score)] = []
                d[str(score)].append(name)
            else:
                d[str(score)].append(name)
for i in d:
    d[i].sort(reverse=True)
d = list(d.items())
d.sort(key=lambda i: int(i[0]), reverse=True)
for key, val in d:
    for i in val:
        print(i, key)

#  Самый дешевый путь
import csv
d = {}
with open('input.csv', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for i in list(reader):
        if i[0].split(';')[-1] == '':
            fr = i[0].split(';')[0]
            too = i[0].split(';')[1]
        else:
            d[i[0].split(';')[0] + i[0].split(';')[1]] = i[0].split(';')[2]
put = fr + ' ' + too
if (fr + too) in d:
    min_tc = int(d[fr + too])
    del d[fr + too]
for kkk in d.keys():
    ts_proezd = 0
    if kkk[0] == fr:
        pause = kkk[1]
        ts_proezd += int(d[kkk])
        for ttt in d.keys():
            if ttt[0] == pause and ttt[1] == too:
                ts_proezd += int(d[ttt])
            if ts_proezd <= min_tc:
                min_tc = ts_proezd
                put = fr + ' ' + pause + ' ' + too
print(put)

#  Первый закон Чизхолма
import csv
data = []
okrug = input()
god1, god2 = input().split()
with open('salary.csv', encoding="utf8") as csvfile:
    reader = csv.DictReader(csvfile, delimiter=';')
    for i in reader:
        if i['Федеральный округ'] == okrug:
            if ((int(i[god2]) - int(i[god1])) / (int(i[god1]) / 100)) < 4:
                k = {}
                k['Субъект'] = i['Субъект']
                k[god1] = i[god1]
                k[god2] = i[god2]
                data.append(k)
with open('out_file.csv', 'w') as f:
    if data:
        writer = csv.DictWriter(f, fieldnames=list(data[0].keys()),
                                delimiter=';', lineterminator='\r\n')
        writer.writeheader()
        for d in data:
            writer.writerow(d)
    else:
        writer = csv.writer(f, delimiter=';')
        writer.writerow('')

#  Экзамены
import csv
import sys
n, m = input().split()
d = []
for line in sys.stdin:
    k = {}
    spis = line.rstrip('\n').split()
    if int(spis[2]) >= int(m) and int(spis[3]) >= int(m) and int(spis[4]) >= int(m):
        if (int(spis[2]) + int(spis[3]) + int(spis[4])) >= int(n):
            k['Фамилия'] = spis[0]
            k['имя'] = spis[1]
            k['результат 1'] = spis[2]
            k['результат 2'] = spis[3]
            k['результат 3'] = spis[4]
            k['сумма'] = (int(spis[2]) + int(spis[3]) + int(spis[4]))
    if k:
        d.append(k)
with open('exam.csv', 'w', encoding='UTF-8') as f:
    writer = csv.DictWriter(f, fieldnames=list(d[0].keys()), delimiter=';')
    writer.writeheader()
    for a in d:
        writer.writerow(a)

#  Перекодировщих
from openpyxl import load_workbook
import csv

wb_val = load_workbook(filename='data.xlsx', data_only=True)
pointSheets = wb_val.sheetnames
sheet_val = wb_val[pointSheets[0]]
with open('output.csv', 'w', newline='') as csvfile:
    writer = csv.writer(
        csvfile, delimiter=";")
    for element in sheet_val:
        res = []
        for el in element:
            res.append(el.value)
        writer.writerow(res)

#  список исполнителей
import sqlite3

a = []
genre = input()
db = sqlite3.connect("music_db.sqlite")
cur = db.cursor()
result = cur.execute(f"""SELECT DISTINCT
  a.name
FROM 
  genre g,
  track t,
  album al,
  artist a
WHERE
  t.genreid = g.genreid 
AND
  t.albumid = al.albumid
AND
  al.artistid = a.artistid
AND
  g.name = '{genre}'
ORDER BY a.name;""")
for i in result:
    if i[0] not in a:
        a.append(i[0])
for i in a:
    print(i)
db.close()
